# accounts/admin_reports_views.py
import csv
import json
from datetime import timedelta, datetime, date

from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from accounts.models import User
from events.models import Event, EventRegistration
from payments.models import PaymentProof
from lostfound.models import LostFoundItem, ClaimRequest


# -----------------------------
# Helpers
# -----------------------------
def _is_admin(user):
    return user.is_authenticated and (getattr(user, "is_superuser", False) or getattr(user, "role", None) == "ADMIN")


def _parse_date(s: str):
    """
    Expect yyyy-mm-dd. Returns date or None.
    """
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def _range_from_request(request):
    """
    Supported:
      ?range=week | month | custom
      custom: ?start=YYYY-MM-DD&end=YYYY-MM-DD
    Returns: (start_dt, end_dt, label)
      end_dt is inclusive (we convert to end_of_day).
    """
    now = timezone.localtime(timezone.now())
    range_key = (request.GET.get("range") or "month").lower().strip()

    if range_key == "week":
        start = (now - timedelta(days=7)).date()
        end = now.date()
        label = "Last 7 days"
    elif range_key == "custom":
        s = _parse_date(request.GET.get("start") or "")
        e = _parse_date(request.GET.get("end") or "")
        if not s or not e:
            # fallback to month if missing
            start = (now - timedelta(days=30)).date()
            end = now.date()
            label = "Last 30 days (fallback)"
        else:
            # keep order
            if e < s:
                s, e = e, s
            start, end = s, e
            label = f"Custom ({start} → {end})"
    else:
        # month default
        start = (now - timedelta(days=30)).date()
        end = now.date()
        label = "Last 30 days"

    # start of day → end of day (inclusive)
    start_dt = timezone.make_aware(datetime.combine(start, datetime.min.time()))
    end_dt = timezone.make_aware(datetime.combine(end, datetime.max.time()))

    return start_dt, end_dt, label, range_key


def _daterange_days(start_dt, end_dt):
    """
    Returns list of ISO date strings between start and end (inclusive).
    """
    start_d = timezone.localtime(start_dt).date()
    end_d = timezone.localtime(end_dt).date()
    days = []
    cur = start_d
    while cur <= end_d:
        days.append(cur.isoformat())
        cur += timedelta(days=1)
    return days


def _safe_dt_field(model, preferred=("created_at", "submitted_at", "registered_at", "updated_at", "date_joined")):
    fields = {f.name for f in model._meta.get_fields() if hasattr(f, "name")}
    for name in preferred:
        if name in fields:
            return name
    return None


# -----------------------------
# Main dashboard
# -----------------------------
@login_required
def admin_reports_view(request):
    if not _is_admin(request.user):
        return HttpResponse("Not allowed", status=403)

    start_dt, end_dt, label, range_key = _range_from_request(request)

    # pick the correct datetime fields safely (you have mixed schemas)
    reg_dt_field = _safe_dt_field(EventRegistration, ("registered_at", "created_at"))
    pay_dt_field = _safe_dt_field(PaymentProof, ("submitted_at", "updated_at", "verified_at", "created_at"))
    claim_dt_field = _safe_dt_field(ClaimRequest, ("created_at",))
    item_dt_field = _safe_dt_field(LostFoundItem, ("created_at",))
    user_dt_field = _safe_dt_field(User, ("date_joined",))

    # ---------------- USERS KPIs ----------------
    users_total = User.objects.count()
    students_total = User.objects.filter(role="STUDENT").count()
    staff_total = User.objects.filter(role="STAFF").count()
    pending_total = User.objects.filter(role__in=["STUDENT", "STAFF"], is_approved=False, is_active=True).count()

    users_in_range = 0
    if user_dt_field:
        users_in_range = User.objects.filter(**{f"{user_dt_field}__range": (start_dt, end_dt)}).count()

    # ---------------- EVENTS KPIs ----------------
    total_events = Event.objects.count()
    published_events = Event.objects.filter(status="published").count()

    regs_total = EventRegistration.objects.count()
    regs_in_range = 0
    if reg_dt_field:
        regs_in_range = EventRegistration.objects.filter(**{f"{reg_dt_field}__range": (start_dt, end_dt)}).count()

    # ---------------- LOST & FOUND KPIs ----------------
    lf_total = LostFoundItem.objects.count()
    lf_open = LostFoundItem.objects.filter(status="open").count()
    lf_returned = LostFoundItem.objects.filter(status="returned").count()

    claims_total = ClaimRequest.objects.count()
    claims_in_range = 0
    if claim_dt_field:
        claims_in_range = ClaimRequest.objects.filter(**{f"{claim_dt_field}__range": (start_dt, end_dt)}).count()

    # ---------------- PAYMENTS KPIs ----------------
    payments_total = PaymentProof.objects.count()
    payments_approved = PaymentProof.objects.filter(status="approved").count()
    payments_rejected = PaymentProof.objects.filter(status="rejected").count()
    payments_pending = PaymentProof.objects.filter(status="pending").count()

    revenue_in_range = 0
    if pay_dt_field:
        revenue_in_range = (
            PaymentProof.objects.filter(status="approved", **{f"{pay_dt_field}__range": (start_dt, end_dt)})
            .aggregate(total=Sum("amount"))["total"]
            or 0
        )

    # ---------------- CHARTS (daily time-series) ----------------
    days = _daterange_days(start_dt, end_dt)

    # registrations per day
    regs_series_map = {}
    if reg_dt_field:
        qs = (
            EventRegistration.objects
            .filter(**{f"{reg_dt_field}__range": (start_dt, end_dt)})
            .annotate(day=TruncDate(reg_dt_field))
            .values("day")
            .annotate(c=Count("id"))
            .order_by("day")
        )
        regs_series_map = {r["day"].isoformat(): int(r["c"]) for r in qs if r["day"]}

    regs_series = [regs_series_map.get(d, 0) for d in days]

    # payments approved sum per day + payments count per day
    pay_count_map = {}
    pay_sum_map = {}
    if pay_dt_field:
        qs_count = (
            PaymentProof.objects
            .filter(**{f"{pay_dt_field}__range": (start_dt, end_dt)})
            .annotate(day=TruncDate(pay_dt_field))
            .values("day")
            .annotate(c=Count("id"))
            .order_by("day")
        )
        pay_count_map = {r["day"].isoformat(): int(r["c"]) for r in qs_count if r["day"]}

        qs_sum = (
            PaymentProof.objects
            .filter(status="approved", **{f"{pay_dt_field}__range": (start_dt, end_dt)})
            .annotate(day=TruncDate(pay_dt_field))
            .values("day")
            .annotate(s=Sum("amount"))
            .order_by("day")
        )
        pay_sum_map = {r["day"].isoformat(): float(r["s"] or 0) for r in qs_sum if r["day"]}

    pay_count_series = [pay_count_map.get(d, 0) for d in days]
    pay_sum_series = [pay_sum_map.get(d, 0) for d in days]

    # ---------------- TABLES (recent activity) ----------------
    recent_regs = (
        EventRegistration.objects
        .select_related("event", "user")
        .order_by("-" + (reg_dt_field or "id"))[:10]
    )

    recent_payments = (
        PaymentProof.objects
        .select_related("registration__event", "registration__user")
        .order_by("-" + (pay_dt_field or "id"))[:10]
    )

    recent_items = (
        LostFoundItem.objects
        .select_related("created_by")
        .order_by("-" + (item_dt_field or "id"))[:10]
    )

    chart_payload = {
        "days": days,
        "registrations": regs_series,
        "payment_count": pay_count_series,
        "payment_approved_sum": pay_sum_series,
        "lf_open": lf_open,
        "lf_returned": lf_returned,
    }

    context = {
        "label": label,
        "range_key": range_key,
        "start": timezone.localtime(start_dt).date().isoformat(),
        "end": timezone.localtime(end_dt).date().isoformat(),

        "users_total": users_total,
        "students_total": students_total,
        "staff_total": staff_total,
        "pending_total": pending_total,
        "users_in_range": users_in_range,

        "total_events": total_events,
        "published_events": published_events,
        "regs_total": regs_total,
        "regs_in_range": regs_in_range,

        "lf_total": lf_total,
        "lf_open": lf_open,
        "lf_returned": lf_returned,
        "claims_total": claims_total,
        "claims_in_range": claims_in_range,

        "payments_total": payments_total,
        "payments_approved": payments_approved,
        "payments_rejected": payments_rejected,
        "payments_pending": payments_pending,
        "revenue_in_range": revenue_in_range,

        "recent_regs": recent_regs,
        "recent_payments": recent_payments,
        "recent_items": recent_items,

        "chart_json": json.dumps(chart_payload),
    }

    return render(request, "accounts/admin_reports.html", context)


# -----------------------------
# EXPORTS (same filters)
# -----------------------------
@login_required
def admin_reports_export_csv(request):
    if not _is_admin(request.user):
        return HttpResponse("Not allowed", status=403)

    start_dt, end_dt, label, range_key = _range_from_request(request)
    reg_dt_field = _safe_dt_field(EventRegistration, ("registered_at", "created_at"))
    pay_dt_field = _safe_dt_field(PaymentProof, ("submitted_at", "updated_at", "verified_at", "created_at"))

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="campus_reports_{range_key}.csv"'

    w = csv.writer(response)
    w.writerow(["Campus Connect Reports", label])
    w.writerow([])
    w.writerow(["Section", "Metric", "Value"])

    # Users
    w.writerow(["Users", "Total users", User.objects.count()])
    w.writerow(["Users", "Students", User.objects.filter(role="STUDENT").count()])
    w.writerow(["Users", "Staff", User.objects.filter(role="STAFF").count()])
    w.writerow(["Users", "Pending approvals", User.objects.filter(role__in=["STUDENT", "STAFF"], is_approved=False, is_active=True).count()])

    # Events
    w.writerow(["Events", "Total events", Event.objects.count()])
    w.writerow(["Events", "Published events", Event.objects.filter(status="published").count()])
    w.writerow(["Events", "Total registrations", EventRegistration.objects.count()])
    if reg_dt_field:
        w.writerow(["Events", "Registrations in range", EventRegistration.objects.filter(**{f"{reg_dt_field}__range": (start_dt, end_dt)}).count()])

    # LostFound
    w.writerow(["LostFound", "Total items", LostFoundItem.objects.count()])
    w.writerow(["LostFound", "Open items", LostFoundItem.objects.filter(status="open").count()])
    w.writerow(["LostFound", "Returned items", LostFoundItem.objects.filter(status="returned").count()])
    w.writerow(["LostFound", "Total claims", ClaimRequest.objects.count()])

    # Payments
    w.writerow(["Payments", "Total payments", PaymentProof.objects.count()])
    w.writerow(["Payments", "Approved", PaymentProof.objects.filter(status="approved").count()])
    w.writerow(["Payments", "Rejected", PaymentProof.objects.filter(status="rejected").count()])
    w.writerow(["Payments", "Pending", PaymentProof.objects.filter(status="pending").count()])
    if pay_dt_field:
        revenue = (
            PaymentProof.objects.filter(status="approved", **{f"{pay_dt_field}__range": (start_dt, end_dt)})
            .aggregate(total=Sum("amount"))["total"]
            or 0
        )
        w.writerow(["Payments", "Revenue in range (NPR)", revenue])

    w.writerow([])
    w.writerow(["Recent Registrations"])
    w.writerow(["Event", "User", "Email"])
    recent_regs = EventRegistration.objects.select_related("event", "user").order_by("-" + (reg_dt_field or "id"))[:25]
    for r in recent_regs:
        w.writerow([getattr(r.event, "title", ""), getattr(r.user, "username", ""), getattr(r.user, "email", "")])

    w.writerow([])
    w.writerow(["Recent Payments"])
    w.writerow(["Event", "User", "Status", "Amount"])
    recent_pay = PaymentProof.objects.select_related("registration__event", "registration__user").order_by("-" + (pay_dt_field or "id"))[:25]
    for p in recent_pay:
        ev = getattr(getattr(p, "registration", None), "event", None)
        us = getattr(getattr(p, "registration", None), "user", None)
        w.writerow([getattr(ev, "title", ""), getattr(us, "username", ""), getattr(p, "status", ""), getattr(p, "amount", "")])

    return response


@login_required
def admin_reports_export_excel(request):
    if not _is_admin(request.user):
        return HttpResponse("Not allowed", status=403)

    # requires openpyxl
    from openpyxl import Workbook

    start_dt, end_dt, label, range_key = _range_from_request(request)
    reg_dt_field = _safe_dt_field(EventRegistration, ("registered_at", "created_at"))
    pay_dt_field = _safe_dt_field(PaymentProof, ("submitted_at", "updated_at", "verified_at", "created_at"))

    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Campus Connect Reports", label])
    ws.append([])
    ws.append(["Section", "Metric", "Value"])

    ws.append(["Users", "Total users", User.objects.count()])
    ws.append(["Users", "Students", User.objects.filter(role="STUDENT").count()])
    ws.append(["Users", "Staff", User.objects.filter(role="STAFF").count()])
    ws.append(["Users", "Pending approvals", User.objects.filter(role__in=["STUDENT", "STAFF"], is_approved=False, is_active=True).count()])

    ws.append(["Events", "Total events", Event.objects.count()])
    ws.append(["Events", "Published events", Event.objects.filter(status="published").count()])
    ws.append(["Events", "Total registrations", EventRegistration.objects.count()])
    if reg_dt_field:
        ws.append(["Events", "Registrations in range", EventRegistration.objects.filter(**{f"{reg_dt_field}__range": (start_dt, end_dt)}).count()])

    ws.append(["LostFound", "Total items", LostFoundItem.objects.count()])
    ws.append(["LostFound", "Open items", LostFoundItem.objects.filter(status="open").count()])
    ws.append(["LostFound", "Returned items", LostFoundItem.objects.filter(status="returned").count()])
    ws.append(["LostFound", "Total claims", ClaimRequest.objects.count()])

    ws.append(["Payments", "Total payments", PaymentProof.objects.count()])
    ws.append(["Payments", "Approved", PaymentProof.objects.filter(status="approved").count()])
    ws.append(["Payments", "Rejected", PaymentProof.objects.filter(status="rejected").count()])
    ws.append(["Payments", "Pending", PaymentProof.objects.filter(status="pending").count()])

    if pay_dt_field:
        revenue = (
            PaymentProof.objects.filter(status="approved", **{f"{pay_dt_field}__range": (start_dt, end_dt)})
            .aggregate(total=Sum("amount"))["total"]
            or 0
        )
        ws.append(["Payments", "Revenue in range (NPR)", float(revenue)])

    # Registrations sheet
    ws2 = wb.create_sheet("Registrations")
    ws2.append(["Event", "User", "Email"])
    recent_regs = EventRegistration.objects.select_related("event", "user").order_by("-" + (reg_dt_field or "id"))[:200]
    for r in recent_regs:
        ws2.append([getattr(r.event, "title", ""), getattr(r.user, "username", ""), getattr(r.user, "email", "")])

    # Payments sheet
    ws3 = wb.create_sheet("Payments")
    ws3.append(["Event", "User", "Status", "Amount", "Currency"])
    recent_pay = PaymentProof.objects.select_related("registration__event", "registration__user").order_by("-" + (pay_dt_field or "id"))[:200]
    for p in recent_pay:
        ev = getattr(getattr(p, "registration", None), "event", None)
        us = getattr(getattr(p, "registration", None), "user", None)
        ws3.append([getattr(ev, "title", ""), getattr(us, "username", ""), getattr(p, "status", ""), float(getattr(p, "amount", 0) or 0), getattr(p, "currency", "")])

    # output response
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="campus_reports_{range_key}.xlsx"'
    wb.save(resp)
    return resp


@login_required
def admin_reports_export_pdf(request):
    if not _is_admin(request.user):
        return HttpResponse("Not allowed", status=403)

    # requires reportlab: pip install reportlab
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    start_dt, end_dt, label, range_key = _range_from_request(request)
    reg_dt_field = _safe_dt_field(EventRegistration, ("registered_at", "created_at"))
    pay_dt_field = _safe_dt_field(PaymentProof, ("submitted_at", "updated_at", "verified_at", "created_at"))

    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="campus_reports_{range_key}.pdf"'

    c = canvas.Canvas(resp, pagesize=A4)
    width, height = A4
    y = height - 50

    def line(text, dy=18, size=11):
        nonlocal y
        c.setFont("Helvetica", size)
        c.drawString(40, y, text)
        y -= dy
        if y < 80:
            c.showPage()
            y = height - 50

    c.setFont("Helvetica-Bold", 16)
    c.drawString(40, y, "Campus Connect — Admin Reports")
    y -= 24
    line(label, dy=20, size=11)
    line("")

    # USERS
    c.setFont("Helvetica-Bold", 12)
    line("Users", dy=18, size=12)
    line(f"Total users: {User.objects.count()}")
    line(f"Students: {User.objects.filter(role='STUDENT').count()}")
    line(f"Staff: {User.objects.filter(role='STAFF').count()}")
    line(f"Pending approvals: {User.objects.filter(role__in=['STUDENT','STAFF'], is_approved=False, is_active=True).count()}")
    line("")

    # EVENTS
    line("Events", dy=18, size=12)
    line(f"Total events: {Event.objects.count()}")
    line(f"Published: {Event.objects.filter(status='published').count()}")
    line(f"Total registrations: {EventRegistration.objects.count()}")
    if reg_dt_field:
        line(f"Registrations in range: {EventRegistration.objects.filter(**{f'{reg_dt_field}__range': (start_dt, end_dt)}).count()}")
    line("")

    # LOSTFOUND
    line("Lost & Found", dy=18, size=12)
    line(f"Total items: {LostFoundItem.objects.count()}")
    line(f"Open items: {LostFoundItem.objects.filter(status='open').count()}")
    line(f"Returned items: {LostFoundItem.objects.filter(status='returned').count()}")
    line(f"Total claims: {ClaimRequest.objects.count()}")
    line("")

    # PAYMENTS
    line("Payments", dy=18, size=12)
    line(f"Total payments: {PaymentProof.objects.count()}")
    line(f"Approved: {PaymentProof.objects.filter(status='approved').count()}")
    line(f"Rejected: {PaymentProof.objects.filter(status='rejected').count()}")
    line(f"Pending: {PaymentProof.objects.filter(status='pending').count()}")
    if pay_dt_field:
        revenue = (
            PaymentProof.objects.filter(status="approved", **{f"{pay_dt_field}__range": (start_dt, end_dt)})
            .aggregate(total=Sum("amount"))["total"]
            or 0
        )
        line(f"Revenue in range (NPR): {revenue}")
    line("")

    # recent registrations (top 10)
    line("Recent Registrations (top 10)", dy=18, size=12)
    recent_regs = EventRegistration.objects.select_related("event", "user").order_by("-" + (reg_dt_field or "id"))[:10]
    for r in recent_regs:
        line(f"- {getattr(r.event,'title','')} | {getattr(r.user,'username','')} | {getattr(r.user,'email','')}", dy=14, size=10)

    c.showPage()
    c.save()
    return resp